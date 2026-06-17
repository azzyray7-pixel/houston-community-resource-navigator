"""
Builds an Excel dashboard from the SQLite database: a raw data sheet, a
neighborhood x category summary matrix (live COUNTIFS formulas), a coverage
gap sheet, and a chart - so the workbook stays dynamic if resources.csv grows.

Run after build_database.py: python scripts/export_dashboard.py
Output: dashboard/Houston_Community_Resource_Dashboard.xlsx
"""
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "database" / "community_resources.db"
OUT_PATH = ROOT / "dashboard" / "Houston_Community_Resource_Dashboard.xlsx"

CATEGORIES = [
    "Refugee & Immigration Services",
    "Multi-Service Community Center",
    "Education & ESL",
    "Healthcare",
    "Food Assistance",
]
FOCUS_NEIGHBORHOODS = ["Gulfton", "Alief", "Spring Branch", "Sharpstown"]

HEADER_FILL = PatternFill("solid", start_color="1F3A5F", end_color="1F3A5F")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F3A5F")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build_dashboard():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    resources = [dict(r) for r in cur.execute("SELECT * FROM resources ORDER BY category, name").fetchall()]
    neighborhoods = [r[0] for r in cur.execute(
        "SELECT DISTINCT neighborhood FROM resources ORDER BY neighborhood"
    ).fetchall()]
    conn.close()

    # Order neighborhoods: focus neighborhoods first, then the rest, Citywide last
    rest = sorted([n for n in neighborhoods if n not in FOCUS_NEIGHBORHOODS and n != "Citywide"])
    ordered_neighborhoods = FOCUS_NEIGHBORHOODS + rest + (["Citywide"] if "Citywide" in neighborhoods else [])

    wb = Workbook()

    # ---------- Sheet 1: Resources ----------
    ws = wb.active
    ws.title = "Resources"
    headers = ["ID", "Name", "Category", "Neighborhood", "ZIP", "Phone", "Website", "Description"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for r in resources:
        ws.append([r["id"], r["name"], r["category"], r["neighborhood"], r["zip"],
                   r["phone"], r["website"], r["description"]])
    last_row = len(resources) + 1
    widths = [5, 34, 26, 16, 8, 14, 24, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=last_row, max_col=len(headers)):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(cell.column == 8))
    ws.freeze_panes = "A2"

    cat_range = f"Resources!$C$2:$C${last_row}"
    neigh_range = f"Resources!$D$2:$D${last_row}"

    # ---------- Sheet 2: Summary matrix ----------
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Resource Count by Neighborhood and Category"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CATEGORIES) + 2)

    header_row = 3
    ws2.cell(row=header_row, column=1, value="Neighborhood")
    for j, cat in enumerate(CATEGORIES, start=2):
        ws2.cell(row=header_row, column=j, value=cat)
    total_col = len(CATEGORIES) + 2
    ws2.cell(row=header_row, column=total_col, value="Total")
    style_header_row(ws2, header_row, total_col)

    first_data_row = header_row + 1
    for i, neighborhood in enumerate(ordered_neighborhoods):
        row = first_data_row + i
        ws2.cell(row=row, column=1, value=neighborhood).font = Font(name="Arial", bold=True, size=10)
        for j, cat in enumerate(CATEGORIES, start=2):
            col_letter = get_column_letter(j)
            formula = f'=COUNTIFS({neigh_range},$A{row},{cat_range},{col_letter}${header_row})'
            ws2.cell(row=row, column=j, value=formula)
        total_letter_start = get_column_letter(2)
        total_letter_end = get_column_letter(len(CATEGORIES) + 1)
        ws2.cell(row=row, column=total_col, value=f"=SUM({total_letter_start}{row}:{total_letter_end}{row})")

    last_data_row = first_data_row + len(ordered_neighborhoods) - 1
    for row in ws2.iter_rows(min_row=first_data_row, max_row=last_data_row, min_col=1, max_col=total_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if cell.column > 1:
                cell.font = BODY_FONT

    # Highlight zero-count cells (gaps) in the category columns
    zero_range = f"B{first_data_row}:{get_column_letter(len(CATEGORIES) + 1)}{last_data_row}"
    ws2.conditional_formatting.add(
        zero_range,
        CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", start_color="FBE2E1", end_color="FBE2E1"))
    )

    ws2.column_dimensions["A"].width = 18
    for j in range(2, total_col + 1):
        ws2.column_dimensions[get_column_letter(j)].width = 16

    # Chart: total resources per neighborhood
    chart = BarChart()
    chart.title = "Total Resources by Neighborhood"
    chart.y_axis.title = "Resource count"
    chart.x_axis.title = "Neighborhood"
    chart.style = 10
    data = Reference(ws2, min_col=total_col, min_row=header_row, max_row=last_data_row)
    cats = Reference(ws2, min_col=1, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 20
    ws2.add_chart(chart, f"A{last_data_row + 3}")

    # ---------- Sheet 3: Coverage Gaps ----------
    ws3 = wb.create_sheet("Coverage Gaps")
    ws3["A1"] = "Coverage Gaps in High Foreign-Born Focus Neighborhoods"
    ws3["A1"].font = TITLE_FONT
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws3["A2"] = "Gulfton, Alief, Spring Branch and Sharpstown are flagged as focus areas based on documented" \
                " high foreign-born / immigrant population share (see data/neighborhood_context.csv)."
    ws3["A2"].font = Font(name="Arial", italic=True, size=9, color="595959")
    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    gap_header_row = 4
    gap_headers = ["Neighborhood", "Category", "Resource Count", "Status"]
    for j, h in enumerate(gap_headers, start=1):
        ws3.cell(row=gap_header_row, column=j, value=h)
    style_header_row(ws3, gap_header_row, len(gap_headers))

    row = gap_header_row + 1
    for neighborhood in FOCUS_NEIGHBORHOODS:
        for cat in CATEGORIES:
            ws3.cell(row=row, column=1, value=neighborhood)
            ws3.cell(row=row, column=2, value=cat)
            ws3.cell(row=row, column=3, value=f'=COUNTIFS({neigh_range},$A{row},{cat_range},$B{row})')
            ws3.cell(row=row, column=4, value=f'=IF(C{row}=0,"Gap - no local provider","Covered")')
            row += 1
    last_gap_row = row - 1

    for r in ws3.iter_rows(min_row=gap_header_row + 1, max_row=last_gap_row, max_col=4):
        for cell in r:
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if cell.column in (3, 4) else "left")

    ws3.conditional_formatting.add(
        f"D{gap_header_row + 1}:D{last_gap_row}",
        CellIsRule(operator="equal", formula=['"Gap - no local provider"'],
                   fill=PatternFill("solid", start_color="FBE2E1", end_color="FBE2E1"),
                   font=Font(name="Arial", size=10, color="A6332A", bold=True))
    )
    ws3.conditional_formatting.add(
        f"D{gap_header_row + 1}:D{last_gap_row}",
        CellIsRule(operator="equal", formula=['"Covered"'],
                   fill=PatternFill("solid", start_color="E2F0D9", end_color="E2F0D9"))
    )
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 24

    OUT_PATH.parent.mkdir(exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Dashboard saved to {OUT_PATH}")


if __name__ == "__main__":
    build_dashboard()
